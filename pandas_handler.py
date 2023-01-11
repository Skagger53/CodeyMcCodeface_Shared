import sys

if __name__ == "__main__":
    print("This is not the main module. Do not execute directly.")
    sys.exit()

import os
import pandas as pd # Using DataFrame, read_excel
import openpyxl as op # Using load_workbook
import numpy as np # Only using NaN to set default values of NaN to ""
import shutil as sh
from error_logger import ErrorLogger

# Manages all dataframes
# Dataframes used for codes imported from PDF and when reading from and writing to Excel
class DataframeHandler:
    def __init__(self,
                 excel_file_path: str,
                 excel_file_sheet_name: str,
                 excel_data_first_row: int,
                 template_file_path: str,
                 log_file_name: str
                 ):
        self.excel_file_path = excel_file_path
        self.excel_file_sheet_name = excel_file_sheet_name
        self.excel_data_first_row = excel_data_first_row # First row of data
        self.template_file_path = template_file_path
        self.header_list = [] # Headers within Excel file

        self.df_excel_import_codes = pd.DataFrame() # Codes imported from Excel
        self.df_pdf_import_codes = pd.DataFrame() # Used for codes imported from PDF

        # Setting up the error logger and messages to the user
        self.error_logger = ErrorLogger(log_file_name)
        self.read_codes_from_excel_errmsg = "An error was encountered reading the Excel data back in.\n\n" \
                                            "Please ensure the Excel file is not corrupt and is properly formatted. " \
                                                "You may need to replace the CM_Codes.xlsx file with the original template file."

    # Called from tkinter_handler.py to create dataframe from codes found in PDF
    def create_df(self, codes_list: list):
        self.df_pdf_import_codes = pd.DataFrame()
        self.df_pdf_import_codes["Codes"] = codes_list

    # Saves codes found from PDF to Excel
    def save_codes_to_excel(self):
        # Error message to the user if the Excel file appears to be open
        save_codes_err_msg = "Failed to save imported codes to Excel file.\n\nPlease close the Excel file if it is open."

        if os.path.isfile(self.excel_file_path): # If file doesn't exist for some reason, don't try to delete it
            try: os.remove(self.excel_file_path)
            except Exception as remove_file_e:
                self.error_logger.log_error("Attempted to remove Excel file before copying template over", remove_file_e)
                return save_codes_err_msg

        try: sh.copy(self.template_file_path, self.excel_file_path)
        except PermissionError: return save_codes_err_msg
        except Exception as copy_file_e:
            self.error_logger.log_error(
                "Attempted to copy template file to be new Excel file for CM use. Not a PermissionError.",
                copy_file_e
            )
            return save_codes_err_msg

        # Loads workbook and relevant sheet name
        wb = op.load_workbook(self.excel_file_path)
        ws = wb[self.excel_file_sheet_name]

        # Writes the data to the table
        for i, code in enumerate(self.df_pdf_import_codes["Codes"]):
            ws["A" + str((i + self.excel_data_first_row))].value = code
        try: wb.save(self.excel_file_path)
        except PermissionError as save_to_excel_e:
            self.error_logger.log_error("Attempting to save imported data to Excel files.", save_to_excel_e)
            return save_codes_err_msg

    # Returns one value: None if successful and string message to user if error encountered
    def read_codes_from_excel(self):
        # ------- Getting accurate header list from Excel file -------
        try:
            codes_wb = op.load_workbook(self.excel_file_path, data_only = True)
            codes_ws = codes_wb[self.excel_file_sheet_name]
        except Exception as load_wb_ws_e:
            self.error_logger.log_error("Attempted to load Codes workbook and worksheet with OpenPyxl", load_wb_ws_e)
            return self.read_codes_from_excel_errmsg

        try:
            self.header_list = []
            # Iterates through 7 columns (A:G) and adds their values to header_list
            for col in codes_ws.iter_cols(min_row = 1, max_row = 1, max_col = 7):
                for cell in col: self.header_list.append(cell.value)
        except Exception as header_setup_e:
            self.error_logger.log_error("Attempted to set up headers with for loop", header_setup_e)
            return self.read_codes_from_excel_errmsg # Same erorr message as above is appropriate

        # ------- Reads data and sets in a dataframe using headers just read in -------
        try:
            converters_dict = {header_name: str for header_name in self.header_list} # Forces string data type
            data = pd.read_excel(
                self.excel_file_path,
                sheet_name = self.excel_file_sheet_name,
                converters = converters_dict
            )
            self.df_excel_import_codes = pd.DataFrame(data, columns = self.header_list)
            self.df_excel_import_codes.replace(np.NaN, "", inplace = True) # Gets rid of Excel's empty cells (NaN)
        except Exception as read_full_excel_data_e:
            self.error_logger.log_error("Attempted to load in all data from Excel file before sending to PCC", read_full_excel_data_e)
            return self.read_codes_from_excel_errmsg